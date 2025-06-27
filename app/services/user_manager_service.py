from sqlalchemy.ext.asyncio import AsyncSession
from ..repositories.user_repository import UserRepository
from ..services.transaction_service import TransactionService
from ..models.models import User, TPointsActivity
from ..core.base import BaseService
from typing import List, Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
import logging
from datetime import datetime, date

logger = logging.getLogger(__name__)

class UserManagerService(BaseService):
    """
    Сервис для управления пользователями через Excel
    - Экспорт/импорт данных пользователей по отделам
    - Начисление/списание T-Points через Excel
    """
    
    def __init__(self, session: AsyncSession, group_management_service=None, bot=None):
        super().__init__(session)
        self.user_repo = UserRepository(session)
        self.transaction_service = TransactionService(session)
        self.group_management_service = group_management_service
        self.bot = bot
    
    async def export_users_to_excel(self) -> BytesIO:
        """Экспорт пользователей в Excel с разделением по отделам"""
        try:
            users = await self.user_repo.get_all_users()
            logger.info(f"Exporting {len(users)} users to Excel")
            
            # Группируем по отделам и статусу
            departments = {}
            users_without_department = []
            inactive_users = []
            
            for user in users:
                if not user.is_active:
                    inactive_users.append(user)
                elif user.department and user.department.strip():
                    dept_name = user.department.strip()
                    if dept_name not in departments:
                        departments[dept_name] = []
                    departments[dept_name].append(user)
                else:
                    users_without_department.append(user)
            
            # Создаем Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # Листы для отделов
                for dept_name, dept_users in departments.items():
                    self._create_users_sheet(writer, dept_users, dept_name)
                
                # Лист для пользователей без отдела
                if users_without_department:
                    self._create_users_sheet(writer, users_without_department, "Сотрудники без отдела")
                
                # Лист для неактивных пользователей
                if inactive_users:
                    self._create_users_sheet(writer, inactive_users, "❌ Неактивные")
                
                # Инструкции
                self._create_users_instructions_sheet(writer)
                
                # Форматирование
                self._format_users_sheets(writer)
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error exporting users: {e}")
            raise
    
    def _create_users_sheet(self, writer, users: List[User], sheet_name: str):
        """Создает лист с пользователями"""
        data = []
        for user in users:
            data.append({
                'telegram_id': user.telegram_id,
                'username': user.username or '',
                'fullname': user.fullname,
                'birth_date': user.birth_date.strftime('%Y-%m-%d') if user.birth_date else '',
                'hire_date': user.hire_date.strftime('%Y-%m-%d') if user.hire_date else '',
                'department': user.department or '',
                'is_active': user.is_active,
                'tpoints': user.tpoints
            })
        
        df = pd.DataFrame(data)
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    
    def _create_users_instructions_sheet(self, writer):
        """Инструкции по работе с пользователями"""
        instructions = [
            ['📋 ИНСТРУКЦИЯ ПО РАБОТЕ С ДАННЫМИ ПОЛЬЗОВАТЕЛЕЙ', ''],
            ['', ''],
            ['🔴 ВАЖНО: НЕ ИЗМЕНЯЙТЕ telegram_id!', 'Это основной ключ'],
            ['', ''],
            ['✅ Что можно изменять:', ''],
            ['- fullname', 'Полное имя'],
            ['- birth_date', 'Дата рождения (YYYY-MM-DD)'],
            ['- hire_date', 'Дата трудоустройства'],
            ['- department', 'Отдел'],
            ['- is_active', 'Активность (TRUE/FALSE)'],
            ['', ''],
            ['🔴 Обязательные поля (красные если пустые):', ''],
            ['- fullname', 'Обязательно'],
            ['', ''],
            ['⚠️ Рекомендуемые поля (предупреждение если пустые):', ''],
            ['- birth_date', 'Рекомендуется для ДР уведомлений'],
            ['', ''],
            ['🚨 ВАЖНО про деактивацию (is_active = FALSE):', ''],
            ['- Пользователь будет АВТОМАТИЧЕСКИ удалён из группы', 'Без уведомления HR!'],
            ['- Пользователь получит уведомление об удалении', 'Автоматически'],
            ['- Данные сохранятся в БД', 'Видно в вкладке "❌ Неактивные"'],
            ['- Можно реактивировать позже', 'is_active = TRUE, но нужно заново пригласить в группу'],
            ['', ''],
            ['⚠️ НЕ изменяйте:', ''],
            ['- telegram_id', 'НЕ ТРОГАТЬ!'],
            ['- username', 'Из Telegram'],
            ['- tpoints', 'Используйте отдельный файл'],
        ]
        
        df = pd.DataFrame(instructions, columns=['Описание', 'Детали'])
        df.to_excel(writer, index=False, sheet_name='📋 ИНСТРУКЦИЯ')
    
    def _format_users_sheets(self, writer):
        """Красное выделение обязательных полей, желтое - рекомендуемых"""
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        for sheet_name, worksheet in writer.sheets.items():
            if '📋 ИНСТРУКЦИЯ' not in sheet_name and '❌ Неактивные' not in sheet_name:
                for row in range(2, worksheet.max_row + 1):
                    # fullname (колонка C) - обязательное поле
                    if not worksheet[f'C{row}'].value:
                        worksheet[f'C{row}'].fill = red_fill
                    
                    # birth_date (колонка D) - рекомендуемое поле
                    if not worksheet[f'D{row}'].value:
                        worksheet[f'D{row}'].fill = yellow_fill
    
    async def export_tpoints_template_to_excel(self) -> BytesIO:
        """Экспорт шаблона для начисления T-Points по отделам"""
        try:
            users = await self.user_repo.get_all_active_users()
            logger.info(f"Creating T-Points template for {len(users)} users")
            
            # Группируем по отделам
            departments = {}
            users_without_department = []
            
            for user in users:
                if user.department and user.department.strip():
                    dept_name = user.department.strip()
                    if dept_name not in departments:
                        departments[dept_name] = []
                    departments[dept_name].append(user)
                else:
                    users_without_department.append(user)
            
            # Создаем Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # Листы для отделов
                for dept_name, dept_users in departments.items():
                    self._create_tpoints_sheet(writer, dept_users, f"{dept_name} - T-Points")
                
                # Лист без отдела
                if users_without_department:
                    self._create_tpoints_sheet(writer, users_without_department, "Без отдела - T-Points")
                
                # Справочники и инструкции
                await self._create_tpoints_activities_sheet(writer)
                self._create_tpoints_instructions_sheet(writer)
                self._create_tpoints_examples_sheet(writer)
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error creating T-Points template: {e}")
            raise
    
    def _create_tpoints_sheet(self, writer, users: List[User], sheet_name: str):
        """Создает лист для начисления T-Points с поддержкой активностей"""
        data = []
        for user in users:
            data.append({
                'telegram_id': user.telegram_id,
                'username': user.username or '',
                'fullname': user.fullname,
                'current_tpoints': user.tpoints,
                'activity_name': '',  # Название активности (для автозаполнения)
                'points_to_add': '',  # Заполняется автоматически или вручную
                'reason': ''  # Заполняется автоматически или вручную
            })
        
        df = pd.DataFrame(data)
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    
    def _create_tpoints_instructions_sheet(self, writer):
        """Инструкции по начислению T-Points"""
        instructions = [
            ['💰 ИНСТРУКЦИЯ ПО НАЧИСЛЕНИЮ T-POINTS', ''],
            ['', ''],
            ['🔴 НЕ ИЗМЕНЯЙТЕ telegram_id, username, fullname!', ''],
            ['', ''],
            ['📋 Порядок заполнения:', ''],
            ['1. Посмотрите лист "🎯 АКТИВНОСТИ"', 'Список доступных активностей'],
            ['2. Заполните одним из способов:', ''],
            ['', ''],
            ['🎯 СПОСОБ 1 - Через активность:', ''],
            ['- activity_name', 'Точное название из листа АКТИВНОСТИ'],
            ['- points_to_add', 'Автоматически (можно изменить)'],
            ['- reason', 'Автоматически (можно изменить)'],
            ['', ''],
            ['✏️ СПОСОБ 2 - Ручное заполнение:', ''],
            ['- activity_name', 'Оставить пустым'],
            ['- points_to_add', '+ для начисления, - для списания'],
            ['- reason', 'Причина операции (обязательно)'],
            ['', ''],
            ['📝 Примеры:', ''],
            ['Хакатон + 100 + "За 1 место"', 'Активность + изменение'],
            ['пусто + 50 + "Бонус к ЗП"', 'Ручное начисление'],
            ['пусто + -25 + "Штраф"', 'Ручное списание'],
            ['пусто + 0 или пусто', 'Пропустить пользователя'],
            ['', ''],
            ['⚠️ Правила:', ''],
            ['- Названия активностей точно как в листе', ''],
            ['- Обязательно указывайте reason', ''],
            ['- Пользователи получат уведомления', ''],
            ['- Отрицательные значения для списания', ''],
        ]
        
        df = pd.DataFrame(instructions, columns=['Описание', 'Детали'])
        df.to_excel(writer, index=False, sheet_name='💰 ИНСТРУКЦИЯ')
    
    def _create_tpoints_examples_sheet(self, writer):
        """Примеры заполнения T-Points"""
        examples = [
            {
                'telegram_id': 123456789,
                'username': 'john_doe',
                'fullname': 'Иванов Иван Иванович',
                'current_tpoints': 150,
                'activity_name': 'Хакатон',
                'points_to_add': 100,
                'reason': 'За 1 место в хакатоне'
            },
            {
                'telegram_id': 987654321,
                'username': 'jane_smith', 
                'fullname': 'Петрова Анна Сергеевна',
                'current_tpoints': 75,
                'activity_name': 'Тимбилдинг',
                'points_to_add': 25,
                'reason': 'Участие в командном квесте'
            },
            {
                'telegram_id': 555666777,
                'username': 'mike_brown',
                'fullname': 'Сидоров Михаил Петрович',
                'current_tpoints': 200,
                'activity_name': '',
                'points_to_add': 50,
                'reason': 'Бонус к зарплате за месяц'
            },
            {
                'telegram_id': 888999000,
                'username': 'anna_white',
                'fullname': 'Кузнецова Анна Александровна',
                'current_tpoints': 120,
                'activity_name': '',
                'points_to_add': -30,
                'reason': 'Штраф за нарушение дресс-кода'
            }
        ]
        
        df = pd.DataFrame(examples)
        df.to_excel(writer, index=False, sheet_name='📝 ПРИМЕРЫ')
    
    async def preview_tpoints_changes(self, file_content: bytes) -> Dict[str, Any]:
        """Предварительный просмотр изменений T-Points"""
        try:
            excel_file = BytesIO(file_content)
            xls = pd.ExcelFile(excel_file)
            
            summary = {
                'total_operations': 0,
                'total_points_add': 0,
                'total_points_remove': 0,
                'operations': [],
                'errors': []
            }
            
            for sheet_name in xls.sheet_names:
                if any(keyword in sheet_name for keyword in ['ИНСТРУКЦИЯ', 'ПРИМЕРЫ', 'АКТИВНОСТИ']):
                    continue
                    
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                for index, row in df.iterrows():
                    try:
                        telegram_id = int(row['telegram_id'])
                        activity_name = row.get('activity_name', '')
                        points_to_add = row.get('points_to_add', '')
                        reason = row.get('reason', '')
                        
                        # Если указана активность, получаем её данные
                        if activity_name and not pd.isna(activity_name) and str(activity_name).strip():
                            activity = await self._get_activity_by_name(str(activity_name).strip())
                            if activity:
                                # Автозаполнение, если поля пустые
                                if pd.isna(points_to_add) or points_to_add == '':
                                    points_to_add = activity.points
                                if pd.isna(reason) or str(reason).strip() == '':
                                    reason = f"Начисление за: {activity.name}"
                            else:
                                summary['errors'].append(f"Строка {index+1} в {sheet_name}: Активность '{activity_name}' не найдена")
                                continue
                        
                        # Пропускаем пустые операции
                        if pd.isna(points_to_add) or points_to_add == '' or points_to_add == 0:
                            continue
                        
                        points_to_add = int(float(points_to_add))
                        
                        if not reason or pd.isna(reason) or str(reason).strip() == '':
                            summary['errors'].append(f"Строка {index+1} в {sheet_name}: нет причины")
                            continue
                        
                        # Проверяем пользователя
                        user = await self.user_repo.get_user_by_telegram_id(telegram_id)
                        if not user:
                            summary['errors'].append(f"Пользователь {telegram_id} не найден")
                            continue
                        
                        operation = {
                            'telegram_id': telegram_id,
                            'username': user.username,
                            'fullname': user.fullname,
                            'current_points': user.tpoints,
                            'points_change': points_to_add,
                            'new_points': user.tpoints + points_to_add,
                            'reason': str(reason).strip(),
                            'activity_name': str(activity_name).strip() if activity_name and not pd.isna(activity_name) else None,
                            'type': 'начисление' if points_to_add > 0 else 'списание'
                        }
                        
                        summary['operations'].append(operation)
                        summary['total_operations'] += 1
                        
                        if points_to_add > 0:
                            summary['total_points_add'] += points_to_add
                        else:
                            summary['total_points_remove'] += abs(points_to_add)
                            
                    except Exception as e:
                        summary['errors'].append(f"Ошибка в строке {index+1}: {str(e)}")
            
            return summary
            
        except Exception as e:
            logger.error(f"Error previewing T-Points: {e}")
            raise
    
    async def apply_tpoints_changes(self, file_content: bytes, bot=None) -> Dict[str, Any]:
        """Применяет изменения T-Points с уведомлениями"""
        try:
            # Предварительный просмотр
            summary = await self.preview_tpoints_changes(file_content)
            
            if not summary['operations']:
                return {
                    'success': True,
                    'message': 'Операций не найдено',
                    'applied': 0
                }
            
            # Применяем операции
            applied = 0
            successful_ops = []
            
            for operation in summary['operations']:
                try:
                    telegram_id = operation['telegram_id']
                    points_change = operation['points_change']
                    reason = operation['reason']
                    
                    # Выполняем транзакцию
                    activity_name = operation.get('activity_name')
                    
                    if activity_name:
                        # Если операция связана с активностью, используем специальный метод
                        activity = await self._get_activity_by_name(activity_name)
                        if activity and points_change > 0:
                            success = await self.transaction_service.create_activity_transaction(
                                user_id=telegram_id,
                                activity_id=activity.id,
                                points=points_change,
                                description=reason
                            )
                        else:
                            # Обычная транзакция, если активность не найдена или это списание
                            if points_change > 0:
                                success = await self.transaction_service.add_points(
                                    user_id=telegram_id,
                                    points=points_change,
                                    description=reason
                                )
                            else:
                                success = await self.transaction_service.remove_points(
                                    user_id=telegram_id,
                                    points=abs(points_change),
                                    description=reason
                                )
                    else:
                        # Ручная операция пополнения/списания
                        if points_change > 0:
                            success = await self.transaction_service.add_points(
                                user_id=telegram_id,
                                points=points_change,
                                description=reason
                            )
                        else:
                            success = await self.transaction_service.remove_points(
                                user_id=telegram_id,
                                points=abs(points_change),
                                description=reason
                            )
                    
                    if success:
                        applied += 1
                        successful_ops.append(operation)
                        
                        # Уведомление пользователю
                        if bot:
                            await self._send_tpoints_notification(bot, operation)
                            
                except Exception as e:
                    logger.error(f"Error applying T-Points for {operation['telegram_id']}: {e}")
            
            return {
                'success': True,
                'message': f'Выполнено {applied} операций T-Points',
                'applied': applied,
                'operations': successful_ops
            }
            
        except Exception as e:
            logger.error(f"Error applying T-Points: {e}")
            return {
                'success': False,
                'message': f'Ошибка: {str(e)}',
                'applied': 0
            }
    
    async def _send_tpoints_notification(self, bot, operation: Dict[str, Any]):
        """Уведомление о начислении/списании T-Points"""
        try:
            telegram_id = operation['telegram_id']
            points_change = operation['points_change']
            reason = operation['reason']
            new_points = operation['new_points']
            
            if points_change > 0:
                emoji = "💰"
                action = "начислено"
                points_text = f"+{points_change}"
            else:
                emoji = "💸"
                action = "списано"
                points_text = f"{points_change}"
            
            message = (
                f"{emoji} T-Points {action}!\n\n"
                f"💎 Изменение: {points_text} T-Points\n"
                f"📝 Причина: {reason}\n"
                f"💰 Текущий баланс: {new_points} T-Points"
            )
            
            await bot.send_message(chat_id=telegram_id, text=message)
            
        except Exception as e:
            logger.error(f"Error sending notification to {operation['telegram_id']}: {e}")
    
    async def _get_activity_by_name(self, activity_name: str) -> Optional['TPointsActivity']:
        """Получить активность по названию"""
        try:
            from ..services.tpoints_activity_service import TPointsActivityService
            activity_service = TPointsActivityService(self.session)
            activities = await activity_service.get_all_activities_full()
            
            for activity in activities:
                if activity.is_active and activity.name.strip().lower() == activity_name.lower():
                    return activity
            return None
        except Exception as e:
            logger.error(f"Error getting activity by name '{activity_name}': {e}")
            return None
    
    async def collect_user_onboarding_data(self, telegram_id: int, fullname: str, 
                                         birth_date: Optional[date] = None,
                                         hire_date: Optional[date] = None,
                                         department: Optional[str] = None) -> bool:
        """Сбор и сохранение данных онбординга нового пользователя"""
        try:
            # Проверяем существование пользователя
            user = await self.user_repo.get_user_by_telegram_id(telegram_id)
            if not user:
                logger.error(f"User {telegram_id} not found for onboarding")
                return False
            
            # Обновляем данные пользователя
            update_data = {}
            
            if birth_date:
                update_data['birth_date'] = birth_date
            
            if hire_date:
                update_data['hire_date'] = hire_date
                
            if department and department.strip():
                update_data['department'] = department.strip()
            
            if update_data:
                await self.user_repo.update_user_data(telegram_id, update_data)
                logger.info(f"Updated onboarding data for user {telegram_id}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error collecting onboarding data for user {telegram_id}: {e}")
            return False

    async def preview_users_import(self, file_content: bytes) -> Dict[str, Any]:
        """Предварительный просмотр изменений данных пользователей"""
        try:
            excel_file = BytesIO(file_content)
            xls = pd.ExcelFile(excel_file)
            
            summary = {
                'total_users': 0,
                'users_to_update': [],
                'errors': [],
                'warnings': []
            }
            
            for sheet_name in xls.sheet_names:
                if '📋 ИНСТРУКЦИЯ' in sheet_name:
                    continue
                    
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                for index, row in df.iterrows():
                    try:
                        telegram_id = int(row['telegram_id'])
                        
                        # Получаем текущего пользователя
                        current_user = await self.user_repo.get_user_by_telegram_id(telegram_id)
                        if not current_user:
                            summary['errors'].append(f"Пользователь {telegram_id} не найден в системе")
                            continue
                        
                        # Анализируем изменения
                        changes = {}
                        warnings = []
                        
                        # Проверяем fullname (обязательное)
                        new_fullname = str(row.get('fullname', '')).strip()
                        if not new_fullname:
                            summary['errors'].append(f"Строка {index+1}: fullname обязательно")
                            continue
                        if new_fullname != current_user.fullname:
                            changes['fullname'] = {
                                'old': current_user.fullname,
                                'new': new_fullname
                            }
                        
                        # Проверяем birth_date (рекомендуемое)
                        birth_date_str = str(row.get('birth_date', '')).strip()
                        if not birth_date_str or birth_date_str == 'nan':
                            warnings.append("Дата рождения не указана - не будет уведомлений о ДР")
                        else:
                            try:
                                new_birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                                current_birth_str = current_user.birth_date.strftime('%Y-%m-%d') if current_user.birth_date else ''
                                if birth_date_str != current_birth_str:
                                    changes['birth_date'] = {
                                        'old': current_birth_str,
                                        'new': birth_date_str
                                    }
                            except ValueError:
                                summary['errors'].append(f"Строка {index+1}: неверный формат birth_date (нужен YYYY-MM-DD)")
                                continue
                        
                        # Проверяем hire_date (опциональное)
                        hire_date_str = str(row.get('hire_date', '')).strip()
                        if hire_date_str and hire_date_str != 'nan':
                            try:
                                new_hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
                                current_hire_str = current_user.hire_date.strftime('%Y-%m-%d') if current_user.hire_date else ''
                                if hire_date_str != current_hire_str:
                                    changes['hire_date'] = {
                                        'old': current_hire_str,
                                        'new': hire_date_str
                                    }
                            except ValueError:
                                summary['errors'].append(f"Строка {index+1}: неверный формат hire_date")
                                continue
                        
                        # Проверяем department
                        new_department = str(row.get('department', '')).strip()
                        current_department = current_user.department or ''
                        if new_department != current_department:
                            changes['department'] = {
                                'old': current_department,
                                'new': new_department
                            }
                        
                        # Проверяем is_active
                        new_is_active = bool(row.get('is_active', True))
                        if new_is_active != current_user.is_active:
                            changes['is_active'] = {
                                'old': current_user.is_active,
                                'new': new_is_active
                            }
                            if not new_is_active:
                                warnings.append("🚨 ДЕАКТИВАЦИЯ: Данный пользователь должен быть удалён из группы!")
                        
                        # Проверка попытки изменить tpoints
                        if 'tpoints' in row and not pd.isna(row['tpoints']):
                            new_tpoints = int(row['tpoints'])
                            if new_tpoints != current_user.tpoints:
                                warnings.append("Изменение T-Points игнорируется (используйте отдельный файл)")
                        
                        # Если есть изменения, добавляем пользователя
                        if changes:
                            user_update = {
                                'telegram_id': telegram_id,
                                'username': current_user.username,
                                'fullname': current_user.fullname,
                                'changes': changes,
                                'warnings': warnings
                            }
                            summary['users_to_update'].append(user_update)
                            summary['total_users'] += 1
                            
                    except Exception as e:
                        summary['errors'].append(f"Ошибка в строке {index+1}: {str(e)}")
            
            return summary
            
        except Exception as e:
            logger.error(f"Error previewing users import: {e}")
            raise
    
    async def import_users_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Импорт обновлений данных пользователей из Excel"""
        try:
            # Предварительный просмотр
            preview = await self.preview_users_import(file_content)
            
            if not preview['users_to_update']:
                return {
                    'success': True,
                    'message': 'Изменений не найдено',
                    'updated': 0
                }
            
            if preview['errors']:
                return {
                    'success': False,
                    'message': f'Найдены ошибки: {len(preview["errors"])}',
                    'errors': preview['errors'],
                    'updated': 0
                }
            
            # Применяем изменения
            updated = 0
            deactivated_users = []
            successful_updates = []
            
            for user_update in preview['users_to_update']:
                try:
                    telegram_id = user_update['telegram_id']
                    changes = user_update['changes']
                    
                    # Подготавливаем данные для обновления
                    update_data = {}
                    
                    for field, change in changes.items():
                        if field == 'birth_date':
                            update_data[field] = datetime.strptime(change['new'], '%Y-%m-%d').date()
                        elif field == 'hire_date' and change['new']:
                            update_data[field] = datetime.strptime(change['new'], '%Y-%m-%d').date()
                        else:
                            update_data[field] = change['new']
                    
                    # Обновляем пользователя
                    success = await self.user_repo.update_user_data(telegram_id, update_data)
                    
                    if success:
                        updated += 1
                        successful_updates.append({
                            'telegram_id': telegram_id,
                            'changes': changes
                        })
                        
                        # Специальное логирование для деактивации
                        if 'is_active' in changes and not changes['is_active']['new']:
                            user = await self.user_repo.get_user_by_telegram_id(telegram_id)
                            logger.warning(f"🚨 USER DEACTIVATED: {user.fullname} (ID: {telegram_id}) - УДАЛЯЕМ ИЗ ГРУППЫ!")
                            deactivated_users.append({
                                'telegram_id': telegram_id,
                                'fullname': user.fullname,
                                'username': user.username
                            })
                            
                            # Автоматическое удаление из группы
                            if self.group_management_service and self.bot:
                                try:
                                    removal_success = await self.group_management_service.remove_user_from_group(
                                        bot=self.bot,
                                        user_id=telegram_id,
                                        reason=f"Деактивация через Excel: {user.fullname}"
                                    )
                                    
                                    if removal_success:
                                        logger.warning(f"✅ USER REMOVED FROM GROUP: {user.fullname} (ID: {telegram_id})")
                                        
                                        # Уведомляем пользователя
                                        await self.group_management_service.notify_user_about_removal(
                                            bot=self.bot,
                                            user_id=telegram_id,
                                            reason="деактивации через HR-систему"
                                        )
                                    else:
                                        logger.error(f"❌ FAILED to remove user {telegram_id} from group")
                                        
                                except Exception as group_error:
                                    logger.error(f"Error removing user {telegram_id} from group: {group_error}")
                            else:
                                logger.warning(f"⚠️ Group management not configured - manual removal required for {telegram_id}")
                        
                        logger.info(f"Updated user {telegram_id}: {list(changes.keys())}")
                        
                except Exception as e:
                    logger.error(f"Error updating user {user_update['telegram_id']}: {e}")
            
            # Формируем сообщение результата
            message = f'Обновлено {updated} пользователей'
            if deactivated_users:
                deactivated_names = [user['fullname'] for user in deactivated_users]
                message += f'\n\n🚨 ДЕАКТИВИРОВАНЫ (удалить из группы): {", ".join(deactivated_names)}'
            
            return {
                'success': True,
                'message': message,
                'updated': updated,
                'deactivated_users': deactivated_users,
                'updates': successful_updates
            }
            
        except Exception as e:
            logger.error(f"Error importing users: {e}")
            return {
                'success': False,
                'message': f'Ошибка импорта: {str(e)}',
                'updated': 0
            }

    async def export_tpoints_journal_to_excel(self, days: int = 30) -> BytesIO:
        """Экспорт журнала T-Points операций в Excel с разделением по отделам"""
        try:
            # Получаем транзакции за период
            from datetime import datetime, timedelta
            since_date = datetime.utcnow() - timedelta(days=days)
            
            # Получаем все транзакции за период
            transactions = await self.transaction_service.billing_repo.get_transactions_since_date(since_date)
            logger.info(f"Exporting {len(transactions)} T-Points transactions to Excel")
            
            if not transactions:
                # Создаем пустой файл с инструкциями
                return await self._create_empty_journal_excel()
            
            # Группируем транзакции по отделам пользователей
            departments = {}
            transactions_without_department = []
            
            for transaction in transactions:
                user = await self.user_repo.get_user_by_telegram_id(transaction.user_id)
                if not user:
                    continue
                    
                if user.department and user.department.strip():
                    dept_name = user.department.strip()
                    if dept_name not in departments:
                        departments[dept_name] = []
                    departments[dept_name].append((transaction, user))
                else:
                    transactions_without_department.append((transaction, user))
            
            # Создаем Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # Листы для отделов
                for dept_name, dept_transactions in departments.items():
                    self._create_journal_sheet(writer, dept_transactions, f"{dept_name} - Операции")
                
                # Лист для пользователей без отдела
                if transactions_without_department:
                    self._create_journal_sheet(writer, transactions_without_department, "Без отдела - Операции")
                
                # Сводный лист
                self._create_journal_summary_sheet(writer, transactions)
                
                # Инструкции
                self._create_journal_instructions_sheet(writer)
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error exporting T-Points journal: {e}")
            raise

    async def _create_empty_journal_excel(self) -> BytesIO:
        """Создает пустой Excel файл журнала с инструкциями"""
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Пустая сводка
            empty_data = pd.DataFrame({
                'Отдел': ['Нет данных'],
                'Операций': [0],
                'Начислено': [0],
                'Списано': [0],
                'Баланс': [0]
            })
            empty_data.to_excel(writer, index=False, sheet_name='📊 СВОДКА')
            
            # Инструкции
            self._create_journal_instructions_sheet(writer)
        
        output.seek(0)
        return output

    def _create_journal_sheet(self, writer, dept_transactions: List[tuple], sheet_name: str):
        """Создает лист с операциями отдела"""
        data = []
        for transaction, user in dept_transactions:
            data.append({
                'ID': transaction.id,
                'Дата': transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Telegram ID': user.telegram_id,
                'Username': user.username or '',
                'ФИО': user.fullname,
                'Отдел': user.department or '',
                'Сумма T-Points': transaction.points_amount,
                'Тип': 'Начисление' if transaction.points_amount > 0 else 'Списание',
                'Описание': transaction.description,
                'Активность': transaction.activity.name if transaction.activity else '',
                'Продукт': transaction.product.name if transaction.product else ''
            })
        
        df = pd.DataFrame(data)
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    def _create_journal_summary_sheet(self, writer, transactions: List):
        """Создает сводный лист со статистикой по отделам"""
        # Группируем по отделам для статистики
        dept_stats = {}
        no_dept_stats = {'operations': 0, 'added': 0, 'removed': 0}
        
        for transaction in transactions:
            # Пользователь уже должен быть загружен с relationship
            user = transaction.user if hasattr(transaction, 'user') and transaction.user else None
                
            if not user:
                continue
                
            department = user.department.strip() if user.department else None
            
            if department:
                if department not in dept_stats:
                    dept_stats[department] = {'operations': 0, 'added': 0, 'removed': 0}
                
                dept_stats[department]['operations'] += 1
                if transaction.points_amount > 0:
                    dept_stats[department]['added'] += transaction.points_amount
                else:
                    dept_stats[department]['removed'] += abs(transaction.points_amount)
            else:
                no_dept_stats['operations'] += 1
                if transaction.points_amount > 0:
                    no_dept_stats['added'] += transaction.points_amount
                else:
                    no_dept_stats['removed'] += abs(transaction.points_amount)
        
        # Формируем данные для сводки
        summary_data = []
        
        for dept_name, stats in dept_stats.items():
            balance = stats['added'] - stats['removed']
            summary_data.append({
                'Отдел': dept_name,
                'Операций': stats['operations'],
                'Начислено T-Points': stats['added'],
                'Списано T-Points': stats['removed'],
                'Баланс отдела': balance
            })
        
        if no_dept_stats['operations'] > 0:
            balance = no_dept_stats['added'] - no_dept_stats['removed']
            summary_data.append({
                'Отдел': 'Без отдела',
                'Операций': no_dept_stats['operations'],
                'Начислено T-Points': no_dept_stats['added'],
                'Списано T-Points': no_dept_stats['removed'],
                'Баланс отдела': balance
            })
        
        # Общая статистика
        total_operations = sum(item['Операций'] for item in summary_data)
        total_added = sum(item['Начислено T-Points'] for item in summary_data)
        total_removed = sum(item['Списано T-Points'] for item in summary_data)
        total_balance = total_added - total_removed
        
        summary_data.append({
            'Отдел': '🔹 ИТОГО',
            'Операций': total_operations,
            'Начислено T-Points': total_added,
            'Списано T-Points': total_removed,
            'Баланс отдела': total_balance
        })
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, index=False, sheet_name='📊 СВОДКА')

    def _create_journal_instructions_sheet(self, writer):
        """Инструкции по журналу операций"""
        instructions = [
            ['📊 ЖУРНАЛ T-POINTS ОПЕРАЦИЙ', ''],
            ['', ''],
            ['📋 Описание листов:', ''],
            ['• 📊 СВОДКА', 'Статистика по отделам'],
            ['• [Отдел] - Операции', 'Детальные операции отдела'],
            ['• Без отдела - Операции', 'Операции пользователей без отдела'],
            ['', ''],
            ['📊 Поля в детальных листах:', ''],
            ['• ID', 'Уникальный номер операции'],
            ['• Дата', 'Дата и время операции'],
            ['• Telegram ID', 'ID пользователя в Telegram'],
            ['• Username', 'Имя пользователя в Telegram'],
            ['• ФИО', 'Полное имя сотрудника'],
            ['• Отдел', 'Отдел сотрудника'],
            ['• Сумма T-Points', 'Количество T-Points (+/-)'],
            ['• Тип', 'Начисление или Списание'],
            ['• Описание', 'Причина операции'],
            ['• Активность', 'Связанная активность (если есть)'],
            ['• Продукт', 'Связанный продукт (если есть)'],
            ['', ''],
            ['💡 Полезные функции Excel:', ''],
            ['• Фильтры', 'Настройте фильтры по датам/типам'],
            ['• Сортировка', 'Сортируйте по любому полю'],
            ['• Поиск', 'Ctrl+F для поиска'],
            ['• Сводные таблицы', 'Создайте свою аналитику'],
        ]
        
        df = pd.DataFrame(instructions, columns=['Описание', 'Детали'])
        df.to_excel(writer, index=False, sheet_name='📋 ИНСТРУКЦИЯ')

    async def _create_tpoints_activities_sheet(self, writer):
        """Создает справочник активностей T-Points"""
        try:
            # Получаем активности из базы данных
            from ..services.tpoints_activity_service import TPointsActivityService
            activity_service = TPointsActivityService(self.session)
            activities = await activity_service.get_all_activities_full()
            
            data = []
            if activities:
                for activity in activities:
                    if activity.is_active:
                        data.append({
                            'Название активности': activity.name,
                            'T-Points': activity.points,
                            'Описание': activity.description or '',
                            'Статус': 'Активна' if activity.is_active else 'Неактивна'
                        })
            
            # Если активностей нет, создаем примеры
            if not data:
                data = [
                    {'Название активности': 'Участие в корпоративе', 'T-Points': 50, 'Описание': 'За участие в корпоративном мероприятии', 'Статус': 'Пример'},
                    {'Название активности': 'Хакатон', 'T-Points': 100, 'Описание': 'За участие в хакатоне', 'Статус': 'Пример'},
                    {'Название активности': 'Обучение', 'T-Points': 30, 'Описание': 'За прохождение обучающего курса', 'Статус': 'Пример'},
                    {'Название активности': 'Тимбилдинг', 'T-Points': 25, 'Описание': 'За участие в командных активностях', 'Статус': 'Пример'},
                    {'Название активности': 'Спорт', 'T-Points': 20, 'Описание': 'За участие в спортивных мероприятиях', 'Статус': 'Пример'}
                ]
            
            df = pd.DataFrame(data)
            df.to_excel(writer, index=False, sheet_name='🎯 АКТИВНОСТИ')
            
        except Exception as e:
            logger.error(f"Error creating activities sheet: {e}")
            # Создаем базовые примеры в случае ошибки
            data = [
                {'Название активности': 'Ошибка загрузки', 'T-Points': 0, 'Описание': 'Не удалось загрузить активности', 'Статус': 'Ошибка'}
            ]
            df = pd.DataFrame(data)
            df.to_excel(writer, index=False, sheet_name='🎯 АКТИВНОСТИ')
