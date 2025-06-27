from sqlalchemy.ext.asyncio import AsyncSession
from ..repositories.catalog_repository import CatalogRepository
from ..core.base import BaseService
import pandas as pd
import openpyxl
import re
import logging
from typing import List, Dict, Any
from io import BytesIO

logger = logging.getLogger(__name__)

class ExcelService(BaseService):
    """
    Сервис для работы с Excel файлами - импорт/экспорт товаров
    ИСПРАВЛЕНО: работает через CatalogRepository
    """
    def __init__(self, session: AsyncSession):
        super().__init__(session)
        self.catalog_repo = CatalogRepository(session)

    def _convert_google_drive_link(self, url: str) -> str:
        """Конвертирует ссылку с Google Drive в прямую ссылку для отображения"""
        if not url or not isinstance(url, str):
            return url
            
        # Паттерн для извлечения ID файла из разных форматов ссылок Google Drive
        patterns = [
            r"https://drive\.google\.com/file/d/([\w-]+)",  # Формат /file/d/
            r"https://drive\.google\.com/open\?id=([\w-]+)",  # Формат ?id=
            r"https://docs\.google\.com/spreadsheets/d/([\w-]+)"  # Формат таблиц
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                file_id = match.group(1)
                # Возвращаем ссылку для просмотра (не для скачивания)
                return f"https://drive.google.com/uc?export=view&id={file_id}"
        
        return url

    def _process_sizes(self, sizes: str) -> Dict[str, int]:
        """Обрабатывает строку с размерами в словарь"""
        if not sizes or pd.isna(sizes):
            return {}
            
        sizes_dict = {}
        try:
            # Проверяем, является ли строка уже словарем
            if isinstance(sizes, dict):
                return sizes
                
            # Удаляем пробелы и разделяем по запятой
            pairs = str(sizes).strip().split(',')
            for pair in pairs:
                if ':' in pair:
                    size, quantity = pair.split(':', 1)  # Ограничиваем split одним разделением
                    sizes_dict[size.strip()] = int(quantity.strip())
        except Exception as e:
            logger.error(f"Error processing sizes '{sizes}': {e}")
            
        return sizes_dict

    async def import_products_from_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Импортирует товары из Excel файла (только локальные файлы)"""
        try:
            # Читаем Excel файл
            df = pd.read_excel(file_path)
            logger.info(f"Excel file loaded: {len(df)} rows")

            products = []
            for index, row in df.iterrows():
                try:
                    # Проверяем обязательные поля
                    if pd.isna(row.get('name')) or pd.isna(row.get('price')):
                        logger.warning(f"Skipping row {index+1}: missing name or price")
                        continue
                    
                    # Конвертируем Google Drive ссылку в image_url если есть
                    image_url = row.get('image_url')
                    if image_url and not pd.isna(image_url):
                        image_url = self._convert_google_drive_link(str(image_url))
                    else:
                        image_url = None
                    
                    # Получаем ID если есть (для обновления существующих товаров)
                    product_id = None
                    if 'id' in row and not pd.isna(row['id']) and str(row['id']).strip():
                        try:
                            product_id = int(row['id'])
                        except (ValueError, TypeError):
                            logger.warning(f"Invalid ID in row {index+1}: {row['id']}")
                    
                    product = {
                        'id': product_id,  # None для новых товаров, ID для обновления
                        'name': str(row['name']).strip(),
                        'description': str(row.get('description', '')).strip() if not pd.isna(row.get('description')) else '',
                        'price': int(float(row['price'])),
                        'image_url': image_url,
                        'color': str(row.get('color', '')).strip() if not pd.isna(row.get('color')) else '',
                        'is_available': bool(row.get('is_available', True)) if not pd.isna(row.get('is_available')) else True
                    }
                    
                    # Обработка размеров или количества для size_quantities
                    if 'sizes' in row and not pd.isna(row['sizes']) and str(row['sizes']).strip():
                        sizes_dict = self._process_sizes(row['sizes'])
                        product['size_quantities'] = sizes_dict  # Словарь размеров
                        product['has_sizes'] = True
                    elif 'quantity' in row and not pd.isna(row['quantity']) and str(row['quantity']).strip():
                        quantity = int(float(row['quantity']))
                        product['size_quantities'] = quantity  # Простое число
                        product['has_sizes'] = False
                    else:
                        product['size_quantities'] = 0
                        product['has_sizes'] = False
                        
                    products.append(product)
                    logger.debug(f"Processed product: {product['name']}")
                    
                except Exception as e:
                    logger.error(f"Error processing row {index+1}: {e}")
                    continue

            logger.info(f"Successfully processed {len(products)} products from Excel")
            return products
            
        except Exception as e:
            logger.error(f"Error importing products from Excel: {e}")
            raise

    async def export_products_to_excel(self) -> BytesIO:
        """Экспортирует товары в Excel файл с шаблоном для заполнения"""
        try:
            # Получаем все товары через репозиторий
            products = await self.catalog_repo.get_all_products()
            
            logger.info(f"Exporting {len(products)} products to Excel")
            
            # Преобразуем текущие товары в формат для DataFrame
            data = []
            for product in products:
                # Формируем строку размеров из size_quantities
                sizes_str = ''
                if product.sizes_dict:  # Если есть размеры в JSON формате
                    sizes_str = ','.join(f"{size}:{qty}" for size, qty in product.sizes_dict.items())
                
                data.append({
                    'id': product.id,
                    'name': product.name,
                    'description': product.description or '',
                    'price': product.price,
                    'image_url': product.image_url or '',
                    'color': product.color or '',
                    'is_available': product.is_available,
                    'sizes': sizes_str,
                    'quantity': product.quantity_as_number if not product.sizes_dict else ''
                })
            
            # Добавляем пустые строки-шаблон для новых товаров
            template_rows = 10  # Количество пустых строк
            for i in range(template_rows):
                data.append({
                    'id': '',
                    'name': '',
                    'description': '',
                    'price': '',
                    'image_url': '',
                    'color': '',
                    'is_available': '',
                    'sizes': '',
                    'quantity': ''
                })
            
            # Создаем DataFrame
            df = pd.DataFrame(data)
            output = BytesIO()
            
            # Используем openpyxl engine для лучшей совместимости
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Основной лист с товарами и шаблоном
                df.to_excel(writer, index=False, sheet_name='Товары')
                
                # Лист с инструкциями
                self._create_instructions_sheet(writer)
                
                # Форматируем основной лист
                self._format_main_sheet(writer.sheets['Товары'], len(products))
            
            output.seek(0)
            logger.info("Excel export completed successfully")
            return output
            
        except Exception as e:
            logger.error(f"Error exporting products to Excel: {e}")
            raise 

    def _create_instructions_sheet(self, writer):
        """Создает лист с инструкциями по заполнению"""
        instructions_data = [
            ['Поле', 'Обязательное', 'Описание', 'Примеры'],
            ['id', 'НЕТ', 'ID товара (автоматический, НЕ ИЗМЕНЯТЬ)', '1, 2, 3, 45'],
            ['name', 'ДА', 'Название товара', 'Футболка Nike, Кроссовки Adidas'],
            ['description', 'НЕТ', 'Описание товара', 'Удобная футболка из хлопка'],
            ['price', 'ДА', 'Цена в T-Points (целое число)', '1500, 2000, 500'],
            ['image_url', 'НЕТ', 'Ссылка на изображение', 'https://example.com/image.jpg'],
            ['color', 'НЕТ', 'Цвет товара', 'Красный, Синий, Черный'],
            ['is_available', 'НЕТ', 'Доступность (true/false)', 'true, false'],
            ['sizes', 'НЕТ', 'Размеры и количество', 'S:10,M:20,L:15,XL:5'],
            ['quantity', 'НЕТ', 'Общее количество (для товаров без размеров)', '100, 50, 25'],
            ['', '', '', ''],
            ['ВАЖНЫЕ ПРАВИЛА:', '', '', ''],
            ['1. Поле ID', '', 'НЕ ИЗМЕНЯЙТЕ ID существующих товаров!', 'Для новых товаров оставьте пустым'],
            ['2. Обязательные поля', '', 'name и price должны быть заполнены', ''],
            ['3. Формат размеров', '', 'Размер:Количество через запятую', 'S:10,M:20,L:15'],
            ['4. Google Drive ссылки', '', 'Автоматически конвертируются в прямые ссылки', ''],
            ['5. Количество', '', 'Если указаны размеры, quantity рассчитается автоматически', ''],
            ['6. Доступность', '', 'По умолчанию true (доступен)', ''],
            ['', '', '', ''],
            ['ТИПЫ ТОВАРОВ:', '', '', ''],
            ['С размерами', '', 'Заполните поле sizes, quantity оставьте пустым', 'S:5,M:10,L:8'],
            ['Без размеров', '', 'Заполните quantity, sizes оставьте пустым', '25'],
            ['', '', '', ''],
            ['ОБНОВЛЕНИЕ ТОВАРОВ:', '', '', ''],
            ['Существующие товары', '', 'Товары с заполненным ID - будут обновлены', 'Измените quantity, price и т.д.'],
            ['Новые товары', '', 'Товары с пустым ID - будут добавлены', 'Заполните name, price обязательно'],
            ['Удаление товаров', '', 'Оставьте поле name пустым', 'Товар не будет обновлен'],
            ['', '', '', ''],
            ['ПРИМЕРЫ ССЫЛОК:', '', '', ''],
            ['Google Drive', '', 'https://drive.google.com/file/d/ID/view', ''],
            ['Прямая ссылка', '', 'https://example.com/image.jpg', ''],
        ]
        
        # Создаем DataFrame для инструкций
        instructions_df = pd.DataFrame(instructions_data[1:], columns=instructions_data[0])
        instructions_df.to_excel(writer, index=False, sheet_name='Инструкции')
        
        # Форматируем лист инструкций
        worksheet = writer.sheets['Инструкции']
        
        # Заголовки жирным шрифтом
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, 5):
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Выделяем важные разделы
        important_font = Font(bold=True, size=11, color="D32F2F")
        for row in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value and ("ПРАВИЛА" in str(cell_value) or "ТИПЫ" in str(cell_value) or "ОБНОВЛЕНИЕ" in str(cell_value) or "ПРИМЕРЫ" in str(cell_value)):
                worksheet.cell(row=row, column=1).font = important_font
        
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _format_main_sheet(self, worksheet, existing_products_count):
        """Форматирует основной лист с товарами и шаблоном"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Стили
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        existing_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Светло-зеленый
        template_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Светло-желтый
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматируем заголовки
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Форматируем существующие товары (зеленый фон)
        for row in range(2, existing_products_count + 2):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = existing_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        
        # Форматируем шаблон для новых товаров (желтый фон)
        for row in range(existing_products_count + 2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = template_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        
        # Добавляем комментарий к первой пустой строке
        if existing_products_count > 0:
            comment_row = existing_products_count + 2
            name_cell = worksheet.cell(row=comment_row, column=1)
            from openpyxl.comments import Comment
            comment = Comment(
                "Пустые строки для добавления новых товаров.\n"
                "Заполните обязательные поля: name и price.\n"
                "Смотрите инструкции на втором листе.",
                "Система"
            )
            name_cell.comment = comment
        
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _create_examples_sheet(self, writer):
        """Удалено - больше не нужен лист с примерами"""
        pass

    def _format_products_sheet(self, worksheet):
        """Переименовано в _format_main_sheet"""
        pass 